"""
สคริปต์นำเข้าข้อมูลรุ่นและยี่ห้อแบตเตอรี่จากไฟล์ Excel เข้าสู่ฐานข้อมูล SQLite (inventory.db)
"""
import os
import re
import sqlite3
import openpyxl
from app_paths import DB_FILE as APP_DB_FILE, resource_path

EXCEL_FILE = str(resource_path("รุ่นแบตเตอรี่.xlsx"))
DB_FILE = str(APP_DB_FILE)

BRANDS = [
    'GLOBAL POWER', 'SCHAEFFLER', 'POWERZONE', 'AMARON', 'BOLIDEN', 
    'PINACO', 'SOLITE', 'BOSCH', 'VARTA', 'YUASA', 'NIKO', 'ULTRA', 
    'ROCKET', 'PUMA', 'LONG', 'BMF', 'ESB', 'FB', 'GS', 'RR'
]

def clean_model_code(raw, brand):
    # ตัดคำว่า แบตเตอรี่, ยี่ห้อ, รุ่น ออก
    text = raw
    text = re.sub(r'^(?:BATTERY|แบตเตอรี่|ยี่ห้อ)\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'(?i)\b' + re.escape(brand) + r'\b', '', text).strip()
    text = re.sub(r'^(?:รุ่น|แบบ)\s*', '', text, flags=re.IGNORECASE).strip()

    # ตัดช่องว่างซ้ำซ้อน
    text = re.sub(r'\s+', ' ', text).strip(' -_/')
    if not text:
        text = raw[:30]
    return text

def import_batteries(excel_path=EXCEL_FILE, db_path=DB_FILE, default_qty=10):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"ไม่พบไฟล์ Excel: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    rows = [str(sheet.cell(row=i, column=1).value).strip() 
            for i in range(2, sheet.max_row + 1) 
            if sheet.cell(row=i, column=1).value]

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    seen_codes = set()
    # ดึง code ที่มีอยู่แล้วใน DB เพื่อไม่ให้ชนกัน
    cursor.execute("SELECT item_code FROM items")
    for row in cursor.fetchall():
        seen_codes.add(row[0])

    imported_count = 0
    updated_count = 0

    for idx, raw in enumerate(rows):
        # ข้ามยี่ห้อ 3K ตามที่ผู้ใช้กำหนด
        if re.search(r'(?i)\b3K\b', raw):
            continue

        # 1. ตรวจหายี่ห้อ
        brand = 'ทั่วไป'
        for b in BRANDS:
            pattern = r'(?i)(?:^|[\s_/\-(])' + re.escape(b) + r'(?:$|[\s_/\-)])'
            if re.search(pattern, raw) or b in raw.upper():
                brand = b
                break

        # 2. ตรวจหาความจุ Ah
        ah_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:Ah|AH|A\b)', raw)
        capacity = f"{ah_match.group(1)} Ah" if ah_match else "-"

        # 3. ตรวจหาประเภทแบตเตอรี่
        raw_u = raw.upper()
        if 'AGM' in raw_u:
            b_type = 'AGM'
        elif 'EFB' in raw_u:
            b_type = 'EFB'
        elif 'SMF' in raw_u or 'แบตแห้ง' in raw or 'แบบแห้ง' in raw or 'แบตเเห้ง' in raw:
            b_type = 'แห้ง (SMF)'
        elif 'AMF' in raw_u:
            b_type = 'AMF (บำรุงรักษาน้อย)'
        elif 'DMF' in raw_u:
            b_type = 'DMF (ไม่ต้องเติมน้ำกลั่น)'
        elif 'น้ำ' in raw:
            b_type = 'น้ำ (Conventional)'
        elif 'HYBRID' in raw_u or 'ไฮบริด' in raw:
            b_type = 'ไฮบริด'
        else:
            b_type = 'กึ่งแห้ง (MF)'

        # 4. สร้างรหัสรุ่น (item_code)
        m_code = clean_model_code(raw, brand)
        if len(m_code) > 40:
            m_code = m_code[:38] + ".."

        base_code = f"{brand} {m_code}".strip() if not m_code.upper().startswith(brand) else m_code
        final_code = base_code
        counter = 1
        while final_code in seen_codes:
            final_code = f"{base_code} (#{counter})"
            counter += 1
        seen_codes.add(final_code)

        # 5. Insert หรือ Update ลงตาราง items
        cursor.execute("""
            INSERT INTO items (brand, item_code, item_name, capacity, battery_type, unit, stock_qty, min_qty)
            VALUES (?, ?, ?, ?, ?, 'ลูก', ?, 3)
            ON CONFLICT(item_code) DO UPDATE SET
                brand=excluded.brand,
                item_name=excluded.item_name,
                capacity=excluded.capacity,
                battery_type=excluded.battery_type
        """, (brand, final_code, raw, capacity, b_type, default_qty))
        imported_count += 1

    conn.commit()

    # นับจำนวนสรุป
    cursor.execute("SELECT COUNT(*), COUNT(DISTINCT brand) FROM items")
    total_items, total_brands = cursor.fetchone()
    conn.close()

    print("=" * 60)
    print(f"✅ นำเข้าข้อมูลแบตเตอรี่สำเร็จ: {imported_count} รายการ")
    print(f"📊 รวมสินค้าในฐานข้อมูล SQLite ปัจจุบัน: {total_items} รุ่น ({total_brands} ยี่ห้อ)")
    print("=" * 60)
    return imported_count, total_items

if __name__ == "__main__":
    import_batteries()
