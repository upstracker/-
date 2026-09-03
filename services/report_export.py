import os
import sys
import subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app_paths import REPORT_DIR

def export_requisitions_to_excel(requisitions, filter_desc="ทั้งหมด", company_name="บริษัท คลังแบตเตอรี่และบริการ จำกัด", output_path=None):
    """
    ส่งออกรายงานสรุปยอดใบเบิกสินค้าเป็นไฟล์ Excel (.xlsx) สวยงาม
    พร้อมหัวตาราง ข้อมูลครบถ้วน และแถวสรุปผลรวมท้ายตาราง
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "สรุปยอดการเบิกจ่าย"

    # ตั้งค่าเส้นตารางให้แสดง
    ws.views.sheetView[0].showGridLines = True

    # สีและฟอนต์มาตรฐาน
    font_title = Font(name="Sarabun", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Sarabun", size=11, bold=True, color="475569")
    font_meta = Font(name="Sarabun", size=10, italic=True, color="64748B")
    
    font_th = Font(name="Sarabun", size=11, bold=True, color="FFFFFF")
    fill_th = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid") # Cyan-800
    
    font_row = Font(name="Sarabun", size=10, color="0F172A")
    font_cancelled = Font(name="Sarabun", size=10, color="DC2626") # แดง
    font_completed = Font(name="Sarabun", size=10, bold=True, color="15803D") # เขียว

    font_total = Font(name="Sarabun", size=11, bold=True, color="0F172A")
    fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    
    double_bottom_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="double", color="0F172A")
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- ส่วนหัวเอกสาร ---
    ws.merge_cells("A1:J1")
    ws["A1"] = company_name
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left

    ws.merge_cells("A2:J2")
    ws["A2"] = f"รายงานสรุปการเบิกจ่ายสินค้าขึ้นรถส่งของ | เงื่อนไข: {filter_desc}"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_left

    ws.merge_cells("A3:J3")
    ws["A3"] = f"พิมพ์รายงานเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | รวมทั้งหมด: {len(requisitions)} ใบเบิก"
    ws["A3"].font = font_meta
    ws["A3"].alignment = align_left

    ws.append([]) # บรรทัดว่าง (Row 4)

    # --- หัวตาราง (Row 5) ---
    headers = [
        "ลำดับ", 
        "เลขที่ใบเบิก", 
        "วันที่เบิก", 
        "ทะเบียนรถ", 
        "คนขับ/ผู้เบิก", 
        "ร้านค้า/ลูกค้า", 
        "สายส่ง/โซน", 
        "บิลอ้างอิง", 
        "จำนวน (ลูก)", 
        "สถานะ"
    ]
    ws.append(headers)
    header_row_idx = 5
    ws.row_dimensions[header_row_idx].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = align_center
        cell.border = thin_border

    # --- ข้อมูลแถว ---
    total_units_sum = 0
    start_data_row = 6

    for idx, r in enumerate(requisitions):
        row_idx = start_data_row + idx
        units = int(r.get("total_units") or 0)
        status_raw = r.get("status", "COMPLETED")
        status_text = "จ่ายของแล้ว" if status_raw == "COMPLETED" else "ยกเลิกแล้ว"
        if status_raw == "COMPLETED":
            total_units_sum += units

        row_data = [
            idx + 1,
            r.get("req_no", "-"),
            r.get("req_date", "-"),
            r.get("vehicle_plate", "-"),
            r.get("driver_name", "-"),
            r.get("customer_name", "-"),
            r.get("route_zone", "-"),
            r.get("ref_bill_no", "-"),
            units,
            status_text
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22

        # กำหนดสไตล์เซลล์ในแถว
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_row
            cell.border = thin_border

            # Alignment
            if col_idx in (1, 2, 3, 4, 8, 10):
                cell.alignment = align_center
            elif col_idx in (5, 6, 7):
                cell.alignment = align_left
            elif col_idx == 9:
                cell.alignment = align_right

            # สีสถานะ
            if col_idx == 10:
                cell.font = font_completed if status_raw == "COMPLETED" else font_cancelled

    # --- แถวสรุปผลรวม (Total Row) ---
    total_row_idx = start_data_row + len(requisitions)
    ws.row_dimensions[total_row_idx].height = 26

    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=8)
    summary_label = ws.cell(row=total_row_idx, column=1)
    summary_label.value = f"รวมใบเบิกที่จ่ายของสำเร็จ ({len([r for r in requisitions if r.get('status') == 'COMPLETED'])} ใบ):"
    summary_label.font = font_total
    summary_label.alignment = align_right
    summary_label.fill = fill_total

    for c in range(1, 9):
        ws.cell(row=total_row_idx, column=c).border = double_bottom_border
        ws.cell(row=total_row_idx, column=c).fill = fill_total

    # ช่องผลรวมจำนวนลูก
    total_units_cell = ws.cell(row=total_row_idx, column=9)
    total_units_cell.value = total_units_sum
    total_units_cell.font = font_total
    total_units_cell.alignment = align_right
    total_units_cell.fill = fill_total
    total_units_cell.border = double_bottom_border

    # ช่องสถานะว่าง
    empty_cell = ws.cell(row=total_row_idx, column=10)
    empty_cell.value = "ลูก"
    empty_cell.font = font_total
    empty_cell.alignment = align_center
    empty_cell.fill = fill_total
    empty_cell.border = double_bottom_border

    # ปรับความกว้างคอลัมน์อัตโนมัติ
    col_widths = {
        1: 8,   # ลำดับ
        2: 20,  # เลขที่ใบเบิก
        3: 14,  # วันที่เบิก
        4: 18,  # ทะเบียนรถ
        5: 22,  # คนขับ
        6: 30,  # ลูกค้า/ร้านค้า
        7: 24,  # สายส่ง/โซน
        8: 16,  # บิลอ้างอิง
        9: 15,  # จำนวน
        10: 16  # สถานะ
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # บันทึกไฟล์
    if not output_path:
        out_dir = str(REPORT_DIR)
        os.makedirs(out_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(out_dir, f"รายงานสรุปการเบิก_{timestamp}.xlsx")

    wb.save(output_path)
    return os.path.abspath(output_path)

def open_exported_excel(filepath):
    """เปิดไฟล์ Excel ด้วยโปรแกรมเปิดสเปรดชีตเริ่มต้นของระบบ"""
    if not os.path.exists(filepath):
        return
    if sys.platform.startswith("win"):
        os.startfile(filepath)
    elif sys.platform.startswith("darwin"):
        subprocess.run(["open", filepath], check=False)
    else:
        subprocess.run(["xdg-open", filepath], check=False)
